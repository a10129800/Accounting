from datetime import datetime
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from .models import Transaction


class ExcelExporter:
    """Export transactions to Excel with formatting and statistics."""

    def __init__(self, output_path: Path | str) -> None:
        self.output_path = Path(output_path)

    def export_transactions(
        self,
        transactions: list[Transaction],
        total_income: int = 0,
        total_expense: int = 0,
        title: str = "交易記錄",
    ) -> bool:
        """Export transactions to Excel with summary.
        
        Args:
            transactions: List of transactions to export
            total_income: Total income amount
            total_expense: Total expense amount
            title: Worksheet title
            
        Returns:
            True if successful, False otherwise
        """
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = title

            # Define styles
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            center_align = Alignment(horizontal="center", vertical="center")
            right_align = Alignment(horizontal="right", vertical="center")
            left_align = Alignment(horizontal="left", vertical="center")
            income_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            expense_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )
            bold_font = Font(bold=True)

            # Title and export time
            ws["A1"] = title
            ws["A1"].font = Font(bold=True, size=14)
            ws["A2"] = f"匯出時間: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
            ws["A2"].font = Font(italic=True, size=10)

            # Headers
            headers = ["日期", "類型", "分類", "金額", "備註"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = center_align
                cell.border = border

            # Data rows
            for row, transaction in enumerate(transactions, 5):
                # Date
                date_cell = ws.cell(row=row, column=1)
                date_cell.value = transaction.transaction_date.isoformat()
                date_cell.alignment = center_align
                date_cell.border = border

                # Type
                type_cell = ws.cell(row=row, column=2)
                type_cell.value = "收入" if transaction.type == "income" else "支出"
                type_cell.alignment = center_align
                type_cell.border = border
                if transaction.type == "income":
                    type_cell.fill = income_fill
                else:
                    type_cell.fill = expense_fill

                # Category
                category_cell = ws.cell(row=row, column=3)
                category_cell.value = transaction.category.name
                category_cell.alignment = left_align
                category_cell.border = border

                # Amount
                amount_cell = ws.cell(row=row, column=4)
                amount_cell.value = transaction.amount
                amount_cell.alignment = right_align
                amount_cell.border = border
                amount_cell.number_format = "#,##0"
                if transaction.type == "income":
                    amount_cell.fill = income_fill
                else:
                    amount_cell.fill = expense_fill

                # Note
                note_cell = ws.cell(row=row, column=5)
                note_cell.value = transaction.note
                note_cell.alignment = left_align
                note_cell.border = border

            # Summary section
            summary_row = len(transactions) + 6
            ws[f"A{summary_row}"] = "摘要"
            ws[f"A{summary_row}"].font = bold_font
            ws[f"A{summary_row}"].fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

            # Summary data
            summary_row += 1
            ws[f"A{summary_row}"] = "總收入:"
            ws[f"A{summary_row}"].font = bold_font
            ws[f"B{summary_row}"] = total_income
            ws[f"B{summary_row}"].font = Font(bold=True, color="008000")
            ws[f"B{summary_row}"].number_format = "#,##0"
            ws[f"B{summary_row}"].alignment = right_align

            summary_row += 1
            ws[f"A{summary_row}"] = "總支出:"
            ws[f"A{summary_row}"].font = bold_font
            ws[f"B{summary_row}"] = total_expense
            ws[f"B{summary_row}"].font = Font(bold=True, color="FF0000")
            ws[f"B{summary_row}"].number_format = "#,##0"
            ws[f"B{summary_row}"].alignment = right_align

            summary_row += 1
            ws[f"A{summary_row}"] = "結餘:"
            ws[f"A{summary_row}"].font = bold_font
            balance = total_income - total_expense
            ws[f"B{summary_row}"] = balance
            ws[f"B{summary_row}"].font = Font(bold=True, color="008000" if balance >= 0 else "FF0000")
            ws[f"B{summary_row}"].number_format = "#,##0"
            ws[f"B{summary_row}"].alignment = right_align

            # Adjust column widths
            ws.column_dimensions["A"].width = 12
            ws.column_dimensions["B"].width = 10
            ws.column_dimensions["C"].width = 12
            ws.column_dimensions["D"].width = 12
            ws.column_dimensions["E"].width = 20

            # Save workbook
            wb.save(self.output_path)
            return True
        except Exception as e:
            print(f"匯出失敗: {e}")
            return False
